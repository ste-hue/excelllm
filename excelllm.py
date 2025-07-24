#!/usr/bin/env python3
"""
ExcelLLM - Convert Excel files to LLM-friendly formats (JSON/Markdown)
Single-file solution for parsing .xlsx files with formulas, values, and metadata.
"""

import json
import argparse
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import zipfile
import xml.etree.ElementTree as ET
from io import BytesIO
import re
from datetime import datetime, date

# Try to import openpyxl for better parsing, fallback to basic XML parsing
try:
    import openpyxl
    from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Using basic XML parsing. Install with: pip install openpyxl")


class ExcelParser:
    """Excel parser that extracts formulas, values, and cell metadata with selective filtering."""

    def __init__(self, chunk_size: int = 200):
        self.chunk_size = chunk_size
        self.sheet_filter = None  # List of sheet names to include
        self.range_filter = None  # Cell range(s) to include (e.g., "A1:C10")

    def parse_file(self, filepath: str, sheets: Optional[List[str]] = None,
                   ranges: Optional[str] = None) -> Dict[str, Any]:
        """Parse Excel file with optional sheet and range filtering.

        Args:
            filepath: Path to Excel file
            sheets: List of sheet names to include (None = all sheets)
            ranges: Cell range(s) to include, e.g. "A1:C10" or "A1:C10,E5:F20"
        """
        self.sheet_filter = sheets
        self.range_filter = ranges

        print(f"[ExcelParser] Parsing with sheet_filter: {sheets}, range_filter: {ranges}")

        if HAS_OPENPYXL:
            return self._parse_with_openpyxl(filepath)
        else:
            return self._parse_with_xml(filepath)

    def _parse_ranges(self, ranges_str: str) -> List[Tuple[int, int, int, int]]:
        """Parse range string into list of (min_col, min_row, max_col, max_row) tuples."""
        if not ranges_str:
            return []

        parsed_ranges = []
        for range_str in ranges_str.split(','):
            range_str = range_str.strip()
            if ':' in range_str:
                # Full range like A1:C10
                bounds = range_boundaries(range_str)
                parsed_ranges.append(bounds)
            else:
                # Single cell like A1
                col_str = ''.join(c for c in range_str if c.isalpha())
                row_str = ''.join(c for c in range_str if c.isdigit())
                col = column_index_from_string(col_str)
                row = int(row_str)
                parsed_ranges.append((col, row, col, row))

        return parsed_ranges

    def _cell_in_ranges(self, row: int, col: int, ranges: List[Tuple[int, int, int, int]]) -> bool:
        """Check if a cell is within any of the specified ranges."""
        if not ranges:
            return True

        for min_col, min_row, max_col, max_row in ranges:
            if min_col <= col <= max_col and min_row <= row <= max_row:
                return True
        return False

    def _parse_with_openpyxl(self, filepath: str) -> Dict[str, Any]:
        """Parse using openpyxl library."""
        wb = openpyxl.load_workbook(filepath, data_only=False)
        result = {
            "filename": Path(filepath).name,
            "sheets": [],
            "filters": {
                "sheets": self.sheet_filter,
                "ranges": self.range_filter
            }
        }

        # Parse ranges once
        parsed_ranges = self._parse_ranges(self.range_filter) if self.range_filter else []
        print(f"[ExcelParser] Parsed ranges: {parsed_ranges}")

        for sheet_name in wb.sheetnames:
            # Skip if sheet filter is set and this sheet isn't included
            if self.sheet_filter and sheet_name not in self.sheet_filter:
                continue

            sheet = wb[sheet_name]
            sheet_data = {
                "name": sheet_name,
                "cells": [],
                "dimensions": f"{sheet.max_row}x{sheet.max_column}",
                "filtered": bool(self.range_filter)
            }

            # Get merged cells info
            merged_map = {}
            for rng in sheet.merged_cells.ranges:
                for cell in rng.cells:
                    merged_map[cell] = str(rng)

            # Extract cells
            row_count = 0
            cells_added = 0

            # If ranges specified, iterate only through those ranges
            if parsed_ranges:
                print(f"[ExcelParser] Processing sheet '{sheet_name}' with range filter")
                for min_col, min_row, max_col, max_row in parsed_ranges:
                    print(f"[ExcelParser] Processing range: col {min_col}-{max_col}, row {min_row}-{max_row}")
                    for row in sheet.iter_rows(min_row=min_row, max_row=max_row,
                                               min_col=min_col, max_col=max_col):
                        if cells_added >= self.chunk_size:
                            sheet_data["truncated"] = True
                            break

                        for cell in row:
                            if cell.value is None and not hasattr(cell, '_value'):
                                continue

                            cell_data = {
                                "address": cell.coordinate,
                                "value": cell.value,
                                "formula": cell.value if cell.data_type == 'f' else None,
                                "type": cell.data_type,
                                "row": cell.row,
                                "col": cell.column
                            }

                            if cell.coordinate in merged_map:
                                cell_data["merged_range"] = merged_map[cell.coordinate]

                            sheet_data["cells"].append(cell_data)
                            cells_added += 1
                print(f"[ExcelParser] Added {cells_added} cells from ranges")
            else:
                # No range filter, process all cells
                print(f"[ExcelParser] Processing sheet '{sheet_name}' without range filter")
                for row in sheet.iter_rows():
                    if cells_added >= self.chunk_size:
                        sheet_data["truncated"] = True
                        sheet_data["total_rows"] = sheet.max_row
                        break

                    for cell in row:
                        if cell.value is None and not hasattr(cell, '_value'):
                            continue

                        cell_data = {
                            "address": cell.coordinate,
                            "value": cell.value,
                            "formula": cell.value if cell.data_type == 'f' else None,
                            "type": cell.data_type,
                            "row": cell.row,
                            "col": cell.column
                        }

                        if cell.coordinate in merged_map:
                            cell_data["merged_range"] = merged_map[cell.coordinate]

                        sheet_data["cells"].append(cell_data)
                        cells_added += 1

                        if cells_added >= self.chunk_size:
                            break

                    row_count += 1

            result["sheets"].append(sheet_data)

        return result

    def _parse_with_xml(self, filepath: str) -> Dict[str, Any]:
        """Basic XML parsing fallback."""
        result = {
            "filename": Path(filepath).name,
            "sheets": [],
            "filters": {
                "sheets": self.sheet_filter,
                "ranges": self.range_filter
            }
        }

        with zipfile.ZipFile(filepath, 'r') as xlsx:
            # Get sheet names
            workbook_xml = xlsx.read('xl/workbook.xml')
            root = ET.fromstring(workbook_xml)

            sheets = []
            for sheet in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                sheets.append({
                    'name': sheet.get('name'),
                    'id': sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                })

            # Parse each sheet
            for idx, sheet_info in enumerate(sheets, 1):
                # Skip if sheet filter is set and this sheet isn't included
                if self.sheet_filter and sheet_info['name'] not in self.sheet_filter:
                    continue

                try:
                    sheet_xml = xlsx.read(f'xl/worksheets/sheet{idx}.xml')
                    sheet_root = ET.fromstring(sheet_xml)

                    sheet_data = {
                        "name": sheet_info['name'],
                        "cells": [],
                        "dimensions": "unknown",
                        "filtered": bool(self.range_filter)
                    }

                    cells_processed = 0
                    for row in sheet_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                        if cells_processed >= self.chunk_size:
                            sheet_data["truncated"] = True
                            break

                        for cell in row.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                            cell_ref = cell.get('r')
                            cell_type = cell.get('t', 'n')

                            # Extract value
                            value_elem = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                            value = value_elem.text if value_elem is not None else None

                            # Extract formula
                            formula_elem = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f')
                            formula = formula_elem.text if formula_elem is not None else None

                            if value or formula:
                                sheet_data["cells"].append({
                                    "address": cell_ref,
                                    "value": value,
                                    "formula": formula,
                                    "type": cell_type
                                })
                                cells_processed += 1

                    result["sheets"].append(sheet_data)

                except Exception as e:
                    print(f"Warning: Could not parse sheet {sheet_info['name']}: {e}")

        return result


class ExcelFormatter:
    """Format parsed Excel data for LLM consumption."""

    @staticmethod
    def to_json(data: Dict[str, Any], pretty: bool = True) -> str:
        """Convert to JSON format."""
        def json_serial(obj):
            """JSON serializer for objects not serializable by default json code"""
            if isinstance(obj, (datetime, date)):
                return obj.isoformat()
            raise TypeError(f"Type {type(obj)} not serializable")

        if pretty:
            return json.dumps(data, indent=2, ensure_ascii=False, default=json_serial)
        return json.dumps(data, ensure_ascii=False, default=json_serial)

    @staticmethod
    def to_markdown(data: Dict[str, Any]) -> str:
        """Convert to Markdown format."""
        lines = [f"# Excel File: {data['filename']}\n"]

        # Add filter info if present
        if data.get('filters'):
            filters = data['filters']
            if filters.get('sheets'):
                lines.append(f"\n**Filtered sheets:** {', '.join(filters['sheets'])}")
            if filters.get('ranges'):
                lines.append(f"**Filtered ranges:** {filters['ranges']}")

        for sheet in data['sheets']:
            lines.append(f"\n## Sheet: {sheet['name']}")
            lines.append(f"Dimensions: {sheet.get('dimensions', 'unknown')}")

            if sheet.get('filtered'):
                lines.append(f"*Note: Showing only cells in specified ranges*")
            elif sheet.get('truncated'):
                lines.append(f"*Note: Showing first {len(sheet['cells'])} cells of {sheet.get('total_rows', 'many')} rows*")

            lines.append("\n| Cell | Value | Formula | Type |")
            lines.append("|------|-------|---------|------|")

            for cell in sheet['cells'][:50]:  # Limit table size for readability
                value = str(cell['value']) if cell['value'] is not None else ''
                formula = cell['formula'] or ''
                cell_type = cell.get('type', '')

                # Escape pipe characters
                value = value.replace('|', '\\|')
                formula = formula.replace('|', '\\|')

                lines.append(f"| {cell['address']} | {value} | {formula} | {cell_type} |")

            if len(sheet['cells']) > 50:
                lines.append(f"\n*... and {len(sheet['cells']) - 50} more cells*")

        return '\n'.join(lines)

    @staticmethod
    def to_simple_text(data: Dict[str, Any]) -> str:
        """Convert to simple text format."""
        lines = [f"Excel File: {data['filename']}"]
        lines.append("=" * 50)

        # Add filter info if present
        if data.get('filters'):
            filters = data['filters']
            if filters.get('sheets'):
                lines.append(f"Filtered sheets: {', '.join(filters['sheets'])}")
            if filters.get('ranges'):
                lines.append(f"Filtered ranges: {filters['ranges']}")
            lines.append("=" * 50)

        for sheet in data['sheets']:
            lines.append(f"\nSheet: {sheet['name']}")
            lines.append("-" * 30)

            for cell in sheet['cells']:
                parts = [f"{cell['address']}:"]
                if cell['value'] is not None:
                    parts.append(f"value={cell['value']}")
                if cell.get('formula'):
                    parts.append(f"formula={cell['formula']}")
                lines.append(" ".join(parts))

        return '\n'.join(lines)


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Convert Excel files to LLM-friendly formats with selective filtering",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  excelllm file.xlsx                    # Output JSON to stdout
  excelllm file.xlsx -o output.json     # Save as JSON
  excelllm file.xlsx -f markdown        # Output as Markdown
  excelllm file.xlsx -f md -o out.md    # Save as Markdown
  excelllm file.xlsx --chunk 500        # Process 500 rows per sheet

  # Selective filtering:
  excelllm file.xlsx -s Sheet1 Sheet2   # Only specific sheets
  excelllm file.xlsx -r A1:C10          # Only specific range
  excelllm file.xlsx -r A1:C10,E5:F20   # Multiple ranges
  excelllm file.xlsx -s Sheet1 -r A1:D5 # Combine sheet and range filters
        """
    )

    parser.add_argument('input', help='Input Excel file (.xlsx)')
    parser.add_argument('-o', '--output', help='Output file (default: stdout)')
    parser.add_argument('-f', '--format',
                       choices=['json', 'markdown', 'md', 'text', 'txt'],
                       default='json',
                       help='Output format (default: json)')
    parser.add_argument('-c', '--chunk',
                       type=int,
                       default=200,
                       help='Max rows per sheet (default: 200)')
    parser.add_argument('--pretty',
                       action='store_true',
                       help='Pretty-print JSON output')
    parser.add_argument('-s', '--sheets',
                       nargs='+',
                       help='Only process specified sheets')
    parser.add_argument('-r', '--ranges',
                       help='Only process specified cell ranges (e.g., A1:C10 or A1:C10,E5:F20)')

    args = parser.parse_args()

    # Validate input file
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: File '{args.input}' not found", file=sys.stderr)
        sys.exit(1)

    if not input_path.suffix.lower() == '.xlsx':
        print(f"Error: Only .xlsx files are supported", file=sys.stderr)
        sys.exit(1)

    # Parse Excel file
    try:
        parser = ExcelParser(chunk_size=args.chunk)
        data = parser.parse_file(args.input, sheets=args.sheets, ranges=args.ranges)
    except Exception as e:
        print(f"Error parsing Excel file: {e}", file=sys.stderr)
        sys.exit(1)

    # Format output
    formatter = ExcelFormatter()

    if args.format in ['json']:
        output = formatter.to_json(data, pretty=args.pretty)
    elif args.format in ['markdown', 'md']:
        output = formatter.to_markdown(data)
    elif args.format in ['text', 'txt']:
        output = formatter.to_simple_text(data)
    else:
        output = formatter.to_json(data, pretty=args.pretty)

    # Write output
    if args.output:
        try:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(output)
            print(f"Output written to: {args.output}")
        except Exception as e:
            print(f"Error writing output file: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        print(output)


if __name__ == '__main__':
    main()
