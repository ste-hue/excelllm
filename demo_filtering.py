#!/usr/bin/env python3
"""
Demo script showing ExcelLLM filtering capabilities in action
Creates a sample Excel file and demonstrates various filtering options
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from excelllm import ExcelParser, ExcelFormatter
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import json

def create_demo_excel():
    """Create a demo Excel file with multiple sheets and data types."""
    wb = openpyxl.Workbook()

    # Sheet 1: Sales Dashboard
    ws1 = wb.active
    ws1.title = "Dashboard"

    # Title
    ws1["A1"] = "Sales Dashboard 2024"
    ws1["A1"].font = Font(size=16, bold=True)

    # KPI Section
    ws1["A3"] = "Key Metrics"
    ws1["A3"].font = Font(bold=True)

    kpis = [
        ("Total Revenue", "B4", 2500000),
        ("Total Costs", "B5", 1800000),
        ("Net Profit", "B6", "=B4-B5"),
        ("Profit Margin", "B7", "=B6/B4"),
        ("YoY Growth", "B8", 0.15)
    ]

    for label, cell, value in kpis:
        ws1[cell[0] + cell[1:]] = label
        ws1[cell] = value

    # Regional Summary (starting at E3)
    ws1["E3"] = "Regional Summary"
    ws1["E3"].font = Font(bold=True)

    regions = ["North", "South", "East", "West"]
    ws1["E4"] = "Region"
    ws1["F4"] = "Sales"
    ws1["G4"] = "Target"
    ws1["H4"] = "Achievement"

    for i, region in enumerate(regions, 5):
        ws1[f"E{i}"] = region
        ws1[f"F{i}"] = 500000 + i * 100000
        ws1[f"G{i}"] = 600000
        ws1[f"H{i}"] = f"=F{i}/G{i}"

    # Sheet 2: Detailed Sales Data
    ws2 = wb.create_sheet("Sales_Data")

    # Headers
    headers = ["Date", "Product", "Region", "Quantity", "Unit Price", "Total", "Commission"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Sample data (100 rows)
    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y"]
    regions = ["North", "South", "East", "West"]

    for row in range(2, 102):
        ws2[f"A{row}"] = f"2024-01-{(row-2) % 30 + 1:02d}"
        ws2[f"B{row}"] = products[(row-2) % 4]
        ws2[f"C{row}"] = regions[(row-2) % 4]
        ws2[f"D{row}"] = 10 + (row % 20)
        ws2[f"E{row}"] = 50 + (row % 10) * 5
        ws2[f"F{row}"] = f"=D{row}*E{row}"
        ws2[f"G{row}"] = f"=F{row}*0.05"

    # Add totals row
    ws2["A103"] = "TOTAL"
    ws2["F103"] = "=SUM(F2:F102)"
    ws2["G103"] = "=SUM(G2:G102)"

    # Sheet 3: Inventory
    ws3 = wb.create_sheet("Inventory")

    ws3["A1"] = "Product"
    ws3["B1"] = "SKU"
    ws3["C1"] = "Stock"
    ws3["D1"] = "Min Level"
    ws3["E1"] = "Status"

    inventory_items = [
        ("Widget A", "WA-001", 150, 50),
        ("Widget B", "WB-001", 30, 50),
        ("Gadget X", "GX-001", 200, 100),
        ("Gadget Y", "GY-001", 45, 75),
    ]

    for i, (product, sku, stock, min_level) in enumerate(inventory_items, 2):
        ws3[f"A{i}"] = product
        ws3[f"B{i}"] = sku
        ws3[f"C{i}"] = stock
        ws3[f"D{i}"] = min_level
        ws3[f"E{i}"] = f'=IF(C{i}<D{i},"REORDER","OK")'

    # Save the file
    wb.save("demo_data.xlsx")
    print("Created demo_data.xlsx")
    return "demo_data.xlsx"


def demonstrate_filtering():
    """Show various filtering examples."""
    filename = create_demo_excel()
    parser = ExcelParser()
    formatter = ExcelFormatter()

    print("\n" + "="*60)
    print("ExcelLLM Filtering Demonstration")
    print("="*60)

    # Example 1: Extract only the Dashboard KPIs
    print("\n1. Extract only KPI section from Dashboard (B4:B8)")
    print("-" * 40)

    data = parser.parse_file(filename, sheets=["Dashboard"], ranges="A4:B8")

    print("Filtered data:")
    for sheet in data['sheets']:
        print(f"\nSheet: {sheet['name']}")
        for cell in sheet['cells']:
            if cell['value'] is not None:
                formula_str = f" [{cell['formula']}]" if cell['formula'] else ""
                print(f"  {cell['address']}: {cell['value']}{formula_str}")

    # Example 2: Extract headers and first 5 rows from Sales_Data
    print("\n\n2. Extract headers and first 5 rows from Sales_Data")
    print("-" * 40)

    data = parser.parse_file(filename, sheets=["Sales_Data"], ranges="A1:G6")

    # Convert to markdown for better visualization
    md_output = formatter.to_markdown(data)
    print(md_output)

    # Example 3: Extract multiple ranges - KPIs and Regional Summary
    print("\n\n3. Extract multiple ranges from Dashboard")
    print("-" * 40)

    data = parser.parse_file(filename, sheets=["Dashboard"], ranges="A3:B8,E3:H8")

    print("Command used: excelllm demo_data.xlsx -s Dashboard -r A3:B8,E3:H8")
    print(f"\nExtracted {len(data['sheets'][0]['cells'])} cells from multiple ranges")

    # Group by range for display
    kpi_cells = [c for c in data['sheets'][0]['cells'] if c['col'] <= 2]
    regional_cells = [c for c in data['sheets'][0]['cells'] if c['col'] >= 5]

    print("\nKPI Section:")
    for cell in kpi_cells:
        if cell['value'] is not None:
            print(f"  {cell['address']}: {cell['value']}")

    print("\nRegional Summary:")
    for cell in regional_cells[:5]:  # Just show first row
        if cell['value'] is not None:
            print(f"  {cell['address']}: {cell['value']}")
    print("  ...")

    # Example 4: Extract only formula cells from all sheets
    print("\n\n4. Extract all formula cells")
    print("-" * 40)

    # Parse entire file
    data = parser.parse_file(filename)

    formula_cells = []
    for sheet in data['sheets']:
        for cell in sheet['cells']:
            if cell.get('formula'):
                formula_cells.append({
                    'sheet': sheet['name'],
                    'address': cell['address'],
                    'formula': cell['formula'],
                    'value': cell['value']
                })

    print(f"Found {len(formula_cells)} formula cells across all sheets:")
    for fc in formula_cells[:10]:  # Show first 10
        print(f"  {fc['sheet']}!{fc['address']}: {fc['formula']} = {fc['value']}")
    if len(formula_cells) > 10:
        print(f"  ... and {len(formula_cells) - 10} more")

    # Example 5: Extract inventory items that need reordering
    print("\n\n5. Extract inventory status column")
    print("-" * 40)

    data = parser.parse_file(filename, sheets=["Inventory"], ranges="A:A,C:E")

    print("Checking inventory status (columns A, C, D, E):")

    # Process the data to find items needing reorder
    inventory_sheet = data['sheets'][0]

    # Group cells by row
    rows = {}
    for cell in inventory_sheet['cells']:
        row = cell['row']
        if row not in rows:
            rows[row] = {}
        rows[row][cell['col']] = cell

    print("\nItems needing attention:")
    for row_num, row_cells in sorted(rows.items()):
        if row_num > 1:  # Skip header
            product = row_cells.get(1, {}).get('value', '')
            stock = row_cells.get(3, {}).get('value', 0)
            min_level = row_cells.get(4, {}).get('value', 0)
            status = row_cells.get(5, {}).get('value', '')

            if status == "REORDER":
                print(f"  ⚠️  {product}: Stock={stock}, Min={min_level}")

    # Example 6: Export filtered data as JSON
    print("\n\n6. Export filtered KPI data as JSON")
    print("-" * 40)

    data = parser.parse_file(filename, sheets=["Dashboard"], ranges="B4:B8")

    # Create a simplified structure for LLM
    kpi_summary = {
        "report": "Sales Dashboard KPIs",
        "metrics": {}
    }

    dashboard = data['sheets'][0]
    for cell in dashboard['cells']:
        if cell['col'] == 1:  # Column A (labels)
            label = cell['value']
            # Find corresponding value in column B
            value_cell = next((c for c in dashboard['cells']
                             if c['row'] == cell['row'] and c['col'] == 2), None)
            if value_cell:
                kpi_summary['metrics'][label] = {
                    'value': value_cell['value'],
                    'formula': value_cell.get('formula')
                }

    print("KPI Summary for LLM:")
    print(json.dumps(kpi_summary, indent=2))

    # Clean up
    print("\n\nDemo complete!")
    print(f"Test file 'demo_data.xlsx' created with 3 sheets")
    print("You can now experiment with different filter combinations!")

    # Show some suggested commands
    print("\n" + "="*60)
    print("Try these commands yourself:")
    print("="*60)
    print("python excelllm.py demo_data.xlsx -s Dashboard -r A1:B8")
    print("python excelllm.py demo_data.xlsx -s Sales_Data -r A1:G10 -f markdown")
    print("python excelllm.py demo_data.xlsx -s Inventory -r E:E")
    print("python excelllm.py demo_data.xlsx -r A1:C1 --pretty")


if __name__ == "__main__":
    demonstrate_filtering()
