#!/usr/bin/env python3
"""
Create test Excel files for testing ExcelLLM
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import random
from datetime import datetime, timedelta

def create_test_excel(filename="test_data.xlsx"):
    """Create a test Excel file with various data types and formulas."""

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Sales Data
    ws1 = wb.active
    ws1.title = "Sales Data"

    # Headers
    headers = ["Product", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Q4 Sales", "Total", "Average", "Growth Rate"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Products
    products = ["Widget A", "Widget B", "Gadget X", "Gadget Y", "Tool Pro", "Tool Lite"]

    # Add data
    for row, product in enumerate(products, 2):
        ws1.cell(row=row, column=1, value=product)

        # Random sales data for Q1-Q4
        for quarter in range(2, 6):
            sales = random.randint(1000, 50000)
            ws1.cell(row=row, column=quarter, value=sales)

        # Total formula
        total_formula = f"=SUM(B{row}:E{row})"
        ws1.cell(row=row, column=6, value=total_formula)

        # Average formula
        avg_formula = f"=AVERAGE(B{row}:E{row})"
        ws1.cell(row=row, column=7, value=avg_formula)

        # Growth rate formula (Q4 vs Q1)
        growth_formula = f"=(E{row}-B{row})/B{row}*100"
        ws1.cell(row=row, column=8, value=growth_formula)

    # Add summary row
    summary_row = len(products) + 2
    ws1.cell(row=summary_row, column=1, value="TOTAL")
    for col in range(2, 8):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{len(products)+1})"
        cell = ws1.cell(row=summary_row, column=col, value=formula)
        cell.font = Font(bold=True)

    # Sheet 2: Financial Model
    ws2 = wb.create_sheet("Financial Model")

    # Basic financial model
    ws2["A1"] = "Financial Projections"
    ws2["A1"].font = Font(size=14, bold=True)

    ws2["A3"] = "Revenue"
    ws2["B3"] = 1000000

    ws2["A4"] = "Growth Rate"
    ws2["B4"] = 0.15

    ws2["A6"] = "Year"
    ws2["B6"] = "Revenue"
    ws2["C6"] = "Costs (70%)"
    ws2["D6"] = "Profit"
    ws2["E6"] = "Margin"

    # 5-year projection
    for year in range(1, 6):
        row = 6 + year
        ws2[f"A{row}"] = 2024 + year - 1

        if year == 1:
            ws2[f"B{row}"] = "=B3"
        else:
            ws2[f"B{row}"] = f"=B{row-1}*(1+$B$4)"

        ws2[f"C{row}"] = f"=B{row}*0.7"
        ws2[f"D{row}"] = f"=B{row}-C{row}"
        ws2[f"E{row}"] = f"=D{row}/B{row}"

    # Sheet 3: Complex Formulas
    ws3 = wb.create_sheet("Complex Formulas")

    # Array-like calculations
    ws3["A1"] = "Matrix Calculations"
    ws3["A1"].font = Font(bold=True)

    # Create a 3x3 matrix
    ws3["A3"] = "Matrix A"
    for i in range(3):
        for j in range(3):
            ws3.cell(row=4+i, column=2+j, value=random.randint(1, 10))

    # Create another 3x3 matrix
    ws3["A8"] = "Matrix B"
    for i in range(3):
        for j in range(3):
            ws3.cell(row=9+i, column=2+j, value=random.randint(1, 10))

    # Sum of matrices
    ws3["A13"] = "Sum (A + B)"
    for i in range(3):
        for j in range(3):
            row = 14 + i
            col = 2 + j
            formula = f"={get_column_letter(col)}{4+i}+{get_column_letter(col)}{9+i}"
            ws3.cell(row=row, column=col, value=formula)

    # Conditional formulas
    ws3["F3"] = "Conditional Tests"
    ws3["F4"] = "Value"
    ws3["G4"] = "Result"

    test_values = [10, 25, 50, 75, 100]
    for i, val in enumerate(test_values):
        row = 5 + i
        ws3[f"F{row}"] = val
        ws3[f"G{row}"] = f'=IF(F{row}>50,"High",IF(F{row}>20,"Medium","Low"))'

    # VLOOKUP example
    ws3["I3"] = "Lookup Table"
    ws3["I4"] = "ID"
    ws3["J4"] = "Name"

    lookup_data = [(1, "Alpha"), (2, "Beta"), (3, "Gamma"), (4, "Delta")]
    for i, (id_val, name) in enumerate(lookup_data):
        row = 5 + i
        ws3[f"I{row}"] = id_val
        ws3[f"J{row}"] = name

    ws3["I10"] = "Lookup ID:"
    ws3["J10"] = 3
    ws3["I11"] = "Result:"
    ws3["J11"] = "=VLOOKUP(J10,I5:J8,2,FALSE)"

    # Sheet 4: Merged Cells
    ws4 = wb.create_sheet("Merged Cells")

    # Title spanning multiple columns
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "Merged Cell Example"
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4["A1"].font = Font(size=16, bold=True)

    # Create a table with merged headers
    ws4.merge_cells("A3:B3")
    ws4["A3"] = "Product Info"
    ws4["A3"].alignment = Alignment(horizontal="center")

    ws4.merge_cells("C3:E3")
    ws4["C3"] = "Sales Data"
    ws4["C3"].alignment = Alignment(horizontal="center")

    # Sub-headers
    ws4["A4"] = "ID"
    ws4["B4"] = "Name"
    ws4["C4"] = "Units"
    ws4["D4"] = "Price"
    ws4["E4"] = "Total"

    # Add some data
    for i in range(5):
        row = 5 + i
        ws4[f"A{row}"] = 100 + i
        ws4[f"B{row}"] = f"Product {chr(65+i)}"
        ws4[f"C{row}"] = random.randint(10, 100)
        ws4[f"D{row}"] = random.randint(50, 200)
        ws4[f"E{row}"] = f"=C{row}*D{row}"

    # Sheet 5: Dates and Times
    ws5 = wb.create_sheet("Dates and Times")

    ws5["A1"] = "Date Operations"
    ws5["A1"].font = Font(bold=True)

    ws5["A3"] = "Start Date"
    ws5["B3"] = datetime(2024, 1, 1)

    ws5["A4"] = "End Date"
    ws5["B4"] = datetime(2024, 12, 31)

    ws5["A5"] = "Days Between"
    ws5["B5"] = "=B4-B3"

    ws5["A7"] = "Date Series"
    base_date = datetime(2024, 1, 1)
    for i in range(12):
        row = 8 + i
        ws5[f"A{row}"] = base_date + timedelta(days=i*30)
        ws5[f"B{row}"] = f"=TEXT(A{row},\"MMMM\")"
        ws5[f"C{row}"] = f"=WEEKDAY(A{row})"

    # Save the workbook
    wb.save(filename)
    print(f"Test Excel file created: {filename}")

    # Create a simple file for quick testing
    wb_simple = openpyxl.Workbook()
    ws = wb_simple.active
    ws.title = "Simple"

    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Double"

    ws["A2"] = "Item 1"
    ws["B2"] = 100
    ws["C2"] = "=B2*2"

    ws["A3"] = "Item 2"
    ws["B3"] = 200
    ws["C3"] = "=B3*2"

    ws["A4"] = "Total"
    ws["B4"] = "=SUM(B2:B3)"
    ws["C4"] = "=SUM(C2:C3)"

    wb_simple.save("simple_test.xlsx")
    print(f"Simple test file created: simple_test.xlsx")


if __name__ == "__main__":
    create_test_excel()
