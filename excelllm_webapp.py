#!/usr/bin/env python3
"""
ExcelLLM Web Application - Modern web interface with drag & drop support
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
from pathlib import Path
import json
import tempfile
import os
import io
from typing import Optional, List

# Import core ExcelLLM functionality
from excelllm import ExcelParser, ExcelFormatter

app = FastAPI(title="ExcelLLM Interactive")

# Enable CORS for web interface
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Store temporary data
temp_files = {}

# HTML Interface with modern UI and drag-drop
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>ExcelLLM Interactive Viewer</title>
    <style>
        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }

        .container {
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            width: 100%;
            max-width: 1200px;
            overflow: hidden;
        }

        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }

        .header h1 {
            font-size: 2.5em;
            margin-bottom: 10px;
        }

        .header p {
            opacity: 0.9;
            font-size: 1.1em;
        }

        .content {
            padding: 40px;
        }

        .step {
            margin-bottom: 40px;
            opacity: 0.3;
            transition: opacity 0.3s;
        }

        .step.active {
            opacity: 1;
        }

        .step-header {
            display: flex;
            align-items: center;
            margin-bottom: 20px;
        }

        .step-number {
            background: #667eea;
            color: white;
            width: 40px;
            height: 40px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: bold;
            margin-right: 15px;
        }

        .step-title {
            font-size: 1.4em;
            color: #333;
        }

        /* Drag & Drop Area */
        .drop-zone {
            border: 3px dashed #667eea;
            border-radius: 15px;
            padding: 60px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s;
            background: #f8f9ff;
        }

        .drop-zone:hover {
            background: #e8ecff;
            border-color: #764ba2;
        }

        .drop-zone.dragover {
            background: #e8ecff;
            border-color: #764ba2;
            transform: scale(1.02);
        }

        .drop-zone .icon {
            font-size: 4em;
            margin-bottom: 20px;
            color: #667eea;
        }

        .drop-zone p {
            color: #666;
            font-size: 1.1em;
            margin-bottom: 10px;
        }

        .file-input {
            display: none;
        }

        .browse-btn {
            background: #667eea;
            color: white;
            border: none;
            padding: 12px 30px;
            border-radius: 8px;
            font-size: 1em;
            cursor: pointer;
            transition: background 0.3s;
        }

        .browse-btn:hover {
            background: #764ba2;
        }

        /* File Info */
        .file-info {
            background: #f8f9ff;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .file-info .name {
            font-weight: bold;
            color: #333;
        }

        .file-info .size {
            color: #666;
            font-size: 0.9em;
        }

        /* Sheet Selection */
        .sheet-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(250px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }

        .sheet-card {
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            padding: 20px;
            cursor: pointer;
            transition: all 0.3s;
            background: white;
        }

        .sheet-card:hover {
            border-color: #667eea;
            box-shadow: 0 5px 15px rgba(102, 126, 234, 0.1);
        }

        .sheet-card.selected {
            background: #f8f9ff;
            border-color: #667eea;
        }

        .sheet-card h3 {
            color: #333;
            margin-bottom: 10px;
        }

        .sheet-card p {
            color: #666;
            font-size: 0.9em;
        }

        /* Range Input */
        .range-input-group {
            display: flex;
            gap: 15px;
            align-items: center;
            margin-bottom: 20px;
        }

        .range-input {
            flex: 1;
            padding: 12px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-size: 1em;
            transition: border-color 0.3s;
        }

        .range-input:focus {
            outline: none;
            border-color: #667eea;
        }

        .range-examples {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            margin-bottom: 20px;
        }

        .range-chip {
            background: #f0f0f0;
            padding: 8px 16px;
            border-radius: 20px;
            font-size: 0.9em;
            color: #666;
            cursor: pointer;
            transition: all 0.3s;
        }

        .range-chip:hover {
            background: #667eea;
            color: white;
        }

        /* Format Selection */
        .format-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }

        .format-card {
            border: 2px solid #e0e0e0;
            border-radius: 10px;
            padding: 20px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s;
        }

        .format-card:hover {
            border-color: #667eea;
        }

        .format-card.selected {
            background: #f8f9ff;
            border-color: #667eea;
        }

        .format-card .icon {
            font-size: 2.5em;
            margin-bottom: 10px;
        }

        .format-card h4 {
            color: #333;
            margin-bottom: 5px;
        }

        .format-card p {
            color: #666;
            font-size: 0.85em;
        }

        /* Process Button */
        .process-btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 15px 50px;
            border-radius: 10px;
            font-size: 1.1em;
            font-weight: bold;
            cursor: pointer;
            transition: transform 0.3s;
            display: block;
            margin: 30px auto;
        }

        .process-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
        }

        .process-btn:disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }

        /* Results */
        .results {
            background: #f8f9ff;
            border-radius: 15px;
            padding: 30px;
            margin-top: 40px;
        }

        .results h2 {
            color: #333;
            margin-bottom: 20px;
        }

        .result-tabs {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
            border-bottom: 2px solid #e0e0e0;
        }

        .tab {
            padding: 10px 20px;
            background: none;
            border: none;
            color: #666;
            font-size: 1em;
            cursor: pointer;
            transition: all 0.3s;
            border-bottom: 3px solid transparent;
            margin-bottom: -2px;
        }

        .tab:hover {
            color: #667eea;
        }

        .tab.active {
            color: #667eea;
            border-bottom-color: #667eea;
        }

        .result-content {
            background: white;
            border-radius: 10px;
            padding: 20px;
            max-height: 500px;
            overflow: auto;
        }

        .result-content pre {
            margin: 0;
            white-space: pre-wrap;
        }

        .result-content table {
            width: 100%;
            border-collapse: collapse;
        }

        .result-content th {
            background: #f0f0f0;
            padding: 12px;
            text-align: left;
            font-weight: 600;
            color: #333;
            position: sticky;
            top: 0;
        }

        .result-content td {
            padding: 10px 12px;
            border-bottom: 1px solid #f0f0f0;
        }

        .result-content tr:hover {
            background: #f8f9ff;
        }

        .download-section {
            display: flex;
            gap: 15px;
            margin-top: 20px;
        }

        .download-btn {
            background: #667eea;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 8px;
            cursor: pointer;
            transition: background 0.3s;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .download-btn:hover {
            background: #764ba2;
        }

        /* Loading */
        .loading {
            display: none;
            text-align: center;
            padding: 40px;
        }

        .spinner {
            border: 4px solid #f3f3f3;
            border-top: 4px solid #667eea;
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }

        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }

        /* Responsive */
        @media (max-width: 768px) {
            .content {
                padding: 20px;
            }

            .header h1 {
                font-size: 2em;
            }

            .sheet-grid {
                grid-template-columns: 1fr;
            }

            .format-grid {
                grid-template-columns: 1fr;
            }

            .range-input-group {
                flex-direction: column;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 ExcelLLM Interactive Viewer</h1>
            <p>Convert Excel files to LLM-ready formats with style</p>
        </div>

        <div class="content">
            <!-- Step 1: File Upload -->
            <div class="step active" id="step1">
                <div class="step-header">
                    <div class="step-number">1</div>
                    <h2 class="step-title">Upload Excel File</h2>
                </div>

                <div class="drop-zone" id="dropZone">
                    <div class="icon">📁</div>
                    <p>Drag and drop your Excel file here</p>
                    <p style="font-size: 0.9em; opacity: 0.7;">or</p>
                    <button class="browse-btn" onclick="document.getElementById('fileInput').click()">
                        Browse Files
                    </button>
                    <input type="file" id="fileInput" class="file-input" accept=".xlsx">
                </div>

                <div id="fileInfo" style="display: none;"></div>
            </div>

            <!-- Step 2: Sheet Selection -->
            <div class="step" id="step2">
                <div class="step-header">
                    <div class="step-number">2</div>
                    <h2 class="step-title">Select Sheets</h2>
                </div>

                <label style="display: flex; align-items: center; margin-bottom: 20px;">
                    <input type="checkbox" id="selectAllSheets" style="margin-right: 10px;">
                    Select all sheets
                </label>

                <div class="sheet-grid" id="sheetGrid"></div>
            </div>

            <!-- Step 3: Range Selection -->
            <div class="step" id="step3">
                <div class="step-header">
                    <div class="step-number">3</div>
                    <h2 class="step-title">Specify Range (Optional)</h2>
                </div>

                <div class="range-input-group">
                    <input type="text" class="range-input" id="rangeInput"
                           placeholder="Enter range (e.g., A1:D10) or leave empty for all">
                </div>

                <div class="range-examples">
                    <div class="range-chip" onclick="setRange('A1:D10')">A1:D10</div>
                    <div class="range-chip" onclick="setRange('A:A')">Column A</div>
                    <div class="range-chip" onclick="setRange('1:10')">Rows 1-10</div>
                    <div class="range-chip" onclick="setRange('A1:C5,E1:G5')">Multiple ranges</div>
                </div>
            </div>

            <!-- Step 4: Output Format -->
            <div class="step" id="step4">
                <div class="step-header">
                    <div class="step-number">4</div>
                    <h2 class="step-title">Choose Output Format</h2>
                </div>

                <div class="format-grid" id="formatGrid">
                    <div class="format-card" data-format="preview">
                        <div class="icon">👁️</div>
                        <h4>Preview</h4>
                        <p>Interactive table view</p>
                    </div>
                    <div class="format-card" data-format="json">
                        <div class="icon">{ }</div>
                        <h4>JSON</h4>
                        <p>Structured data</p>
                    </div>
                    <div class="format-card" data-format="markdown">
                        <div class="icon">📝</div>
                        <h4>Markdown</h4>
                        <p>Formatted text</p>
                    </div>
                    <div class="format-card" data-format="text">
                        <div class="icon">📄</div>
                        <h4>Plain Text</h4>
                        <p>Simple format</p>
                    </div>
                </div>

                <button class="process-btn" id="processBtn" onclick="processFile()">
                    🚀 Process Excel
                </button>
            </div>

            <!-- Loading -->
            <div class="loading" id="loading">
                <div class="spinner"></div>
                <p>Processing your file...</p>
            </div>

            <!-- Results -->
            <div class="results" id="results" style="display: none;">
                <h2>Results</h2>

                <div class="result-tabs" id="resultTabs"></div>
                <div class="result-content" id="resultContent"></div>

                <div class="download-section" id="downloadSection"></div>
            </div>
        </div>
    </div>

    <script>
        let uploadedFile = null;
        let fileData = null;
        let selectedSheets = new Set();
        let selectedFormat = 'preview';

        // Initialize
        document.addEventListener('DOMContentLoaded', function() {
            setupDragDrop();
            setupFileInput();
            setupFormatSelection();
            setupSheetSelection();
        });

        // Drag and Drop
        function setupDragDrop() {
            const dropZone = document.getElementById('dropZone');

            dropZone.addEventListener('dragover', (e) => {
                e.preventDefault();
                dropZone.classList.add('dragover');
            });

            dropZone.addEventListener('dragleave', () => {
                dropZone.classList.remove('dragover');
            });

            dropZone.addEventListener('drop', (e) => {
                e.preventDefault();
                dropZone.classList.remove('dragover');

                const files = e.dataTransfer.files;
                if (files.length > 0 && files[0].name.endsWith('.xlsx')) {
                    handleFile(files[0]);
                } else {
                    alert('Please drop an Excel (.xlsx) file');
                }
            });
        }

        // File Input
        function setupFileInput() {
            const fileInput = document.getElementById('fileInput');
            fileInput.addEventListener('change', (e) => {
                if (e.target.files.length > 0) {
                    handleFile(e.target.files[0]);
                }
            });
        }

        // Handle File Upload
        async function handleFile(file) {
            uploadedFile = file;

            // Show file info
            const fileInfo = document.getElementById('fileInfo');
            fileInfo.innerHTML = `
                <div class="file-info">
                    <div>
                        <div class="name">📄 ${file.name}</div>
                        <div class="size">${(file.size / 1024).toFixed(1)} KB</div>
                    </div>
                    <button class="browse-btn" onclick="document.getElementById('fileInput').click()">
                        Change File
                    </button>
                </div>
            `;
            fileInfo.style.display = 'block';

            // Upload file and get sheet info
            const formData = new FormData();
            formData.append('file', file);

            try {
                const response = await fetch('/api/upload', {
                    method: 'POST',
                    body: formData
                });

                if (!response.ok) throw new Error('Upload failed');

                fileData = await response.json();
                displaySheets(fileData.sheets);

                // Activate next steps
                document.getElementById('step2').classList.add('active');
                document.getElementById('step3').classList.add('active');
                document.getElementById('step4').classList.add('active');

            } catch (error) {
                alert('Error uploading file: ' + error.message);
            }
        }

        // Display Sheet Selection
        function displaySheets(sheets) {
            const grid = document.getElementById('sheetGrid');
            grid.innerHTML = '';

            sheets.forEach((sheet, index) => {
                const card = document.createElement('div');
                card.className = 'sheet-card';
                card.dataset.sheet = sheet.name;
                card.innerHTML = `
                    <h3>${sheet.name}</h3>
                    <p>${sheet.rows} rows × ${sheet.cols} columns</p>
                `;

                card.addEventListener('click', () => {
                    if (selectedSheets.has(sheet.name)) {
                        selectedSheets.delete(sheet.name);
                        card.classList.remove('selected');
                    } else {
                        selectedSheets.add(sheet.name);
                        card.classList.add('selected');
                    }
                    updateSelectAll();
                });

                grid.appendChild(card);
            });
        }

        // Sheet Selection
        function setupSheetSelection() {
            const selectAll = document.getElementById('selectAllSheets');
            selectAll.addEventListener('change', (e) => {
                const cards = document.querySelectorAll('.sheet-card');
                if (e.target.checked) {
                    cards.forEach(card => {
                        selectedSheets.add(card.dataset.sheet);
                        card.classList.add('selected');
                    });
                } else {
                    selectedSheets.clear();
                    cards.forEach(card => card.classList.remove('selected'));
                }
            });
        }

        function updateSelectAll() {
            const selectAll = document.getElementById('selectAllSheets');
            const totalSheets = document.querySelectorAll('.sheet-card').length;
            selectAll.checked = selectedSheets.size === totalSheets;
        }

        // Range Input
        function setRange(range) {
            document.getElementById('rangeInput').value = range;
        }

        // Format Selection
        function setupFormatSelection() {
            const cards = document.querySelectorAll('.format-card');
            cards.forEach(card => {
                card.addEventListener('click', () => {
                    // Remove previous selection
                    cards.forEach(c => c.classList.remove('selected'));
                    // Add selection
                    card.classList.add('selected');
                    selectedFormat = card.dataset.format;
                });
            });

            // Select preview by default
            document.querySelector('[data-format="preview"]').classList.add('selected');
        }

        // Process File
        async function processFile() {
            if (!uploadedFile) {
                alert('Please upload a file first');
                return;
            }

            // Show loading
            document.getElementById('loading').style.display = 'block';
            document.getElementById('results').style.display = 'none';

            // Prepare request
            const formData = new FormData();
            formData.append('file', uploadedFile);
            formData.append('sheets', Array.from(selectedSheets).join(','));
            formData.append('range', document.getElementById('rangeInput').value);
            formData.append('format', selectedFormat);

            try {
                const response = await fetch('/api/process', {
                    method: 'POST',
                    body: formData
                });

                if (!response.ok) throw new Error('Processing failed');

                const result = await response.json();
                window.currentResult = result;  // Store for tab switching

                // Check if the result is large and warn user
                if (result.data && result.data.sheets) {
                    let totalCells = 0;
                    result.data.sheets.forEach(sheet => {
                        totalCells += (sheet.cells || []).length;
                    });
                    if (totalCells > 10000) {
                        console.warn(`Large dataset: ${totalCells} cells`);
                        alert(`Warning: This file contains ${totalCells.toLocaleString()} cells. Downloads may take longer than usual.`);
                    }
                } else if (result.total_cells && result.total_cells > 10000) {
                    console.warn(`Large dataset: ${result.total_cells} cells`);
                    alert(`Warning: This file contains ${result.total_cells.toLocaleString()} cells. Downloads may take longer than usual.`);
                }

                displayResults(result);

            } catch (error) {
                alert('Error processing file: ' + error.message);
            } finally {
                document.getElementById('loading').style.display = 'none';
            }
        }

        // Display Results
        function displayResults(result) {
            const resultsDiv = document.getElementById('results');
            const tabs = document.getElementById('resultTabs');
            const content = document.getElementById('resultContent');
            const downloads = document.getElementById('downloadSection');

            // Clear previous results
            tabs.innerHTML = '';
            content.innerHTML = '';
            downloads.innerHTML = '';

            // Create tabs for each sheet
            if (result.format === 'preview') {
                result.data.sheets.forEach((sheet, index) => {
                    const tab = document.createElement('button');
                    tab.className = 'tab' + (index === 0 ? ' active' : '');
                    tab.textContent = sheet.name;
                    tab.onclick = () => showSheetTab(index);
                    tabs.appendChild(tab);
                });

                // Show first sheet
                showSheetData(result.data.sheets[0]);

            } else {
                // Single tab for other formats
                const tab = document.createElement('button');
                tab.className = 'tab active';
                tab.textContent = result.format.toUpperCase();
                tabs.appendChild(tab);

                // Display formatted content
                if (result.format === 'json') {
                    content.innerHTML = `<pre>${JSON.stringify(result.data, null, 2)}</pre>`;
                } else {
                    content.innerHTML = `<pre>${result.content}</pre>`;
                }
            }

            // Add download buttons
            const formats = ['json', 'markdown', 'text', 'csv'];
            formats.forEach(fmt => {
                const btn = document.createElement('button');
                btn.className = 'download-btn';
                btn.innerHTML = `📥 Download as ${fmt.toUpperCase()}`;
                btn.onclick = () => downloadAs(fmt);
                downloads.appendChild(btn);
            });

            resultsDiv.style.display = 'block';
            resultsDiv.scrollIntoView({ behavior: 'smooth', block: 'nearest' });
        }

        function showSheetTab(index) {
            // Update tab states
            const tabs = document.querySelectorAll('.tab');
            tabs.forEach((tab, i) => {
                tab.classList.toggle('active', i === index);
            });

            // Show sheet data
            if (window.currentResult && window.currentResult.data && window.currentResult.data.sheets[index]) {
                showSheetData(window.currentResult.data.sheets[index]);
            }
        }

        function showSheetData(sheet) {
            const content = document.getElementById('resultContent');

            // Create table
            let html = '<table><thead><tr>';
            html += '<th>Cell</th><th>Value</th><th>Type</th><th>Formula</th>';
            html += '</tr></thead><tbody>';

            sheet.cells.forEach(cell => {
                html += '<tr>';
                html += `<td>${cell.address}</td>`;
                html += `<td>${cell.value || ''}</td>`;
                html += `<td>${cell.type}</td>`;
                html += `<td>${cell.formula || ''}</td>`;
                html += '</tr>';
            });

            html += '</tbody></table>';
            content.innerHTML = html;
        }

        async function downloadAs(format) {
            const formData = new FormData();
            formData.append('file', uploadedFile);
            formData.append('sheets', Array.from(selectedSheets).join(','));
            formData.append('range', document.getElementById('rangeInput').value);
            formData.append('format', format);

            try {
                // Show loading indicator
                const originalText = event.target.innerText;
                event.target.innerText = 'Downloading...';
                event.target.disabled = true;

                // Create a direct download URL using form submission
                const response = await fetch('/api/download', {
                    method: 'POST',
                    body: formData
                });

                if (!response.ok) {
                    const errorText = await response.text();
                    console.error('Download error:', errorText);
                    throw new Error(`Download failed: ${response.status} - ${errorText}`);
                }

                // Get filename from Content-Disposition header
                const contentDisposition = response.headers.get('content-disposition');
                let filename = `output.${format}`;
                if (contentDisposition) {
                    const matches = /filename="(.+)"/.exec(contentDisposition);
                    if (matches && matches[1]) {
                        filename = matches[1];
                    }
                }

                // Alternative download method for better compatibility
                const blob = await response.blob();

                // Use different download method based on browser
                if (window.navigator && window.navigator.msSaveOrOpenBlob) {
                    // IE/Edge Legacy
                    window.navigator.msSaveOrOpenBlob(blob, filename);
                } else {
                    // Modern browsers
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.style.display = 'none';
                    a.href = url;
                    a.download = filename;
                    document.body.appendChild(a);
                    a.click();

                    // Cleanup
                    window.URL.revokeObjectURL(url);
                    document.body.removeChild(a);
                }

                // Restore button
                event.target.innerText = originalText;
                event.target.disabled = false;

            } catch (error) {
                console.error('Full download error:', error);
                alert('Error downloading file: ' + error.message);
                // Restore button on error
                if (event.target) {
                    event.target.innerText = originalText || 'Download';
                    event.target.disabled = false;
                }
            }
        }
    </script>
</body>
</html>
"""

@app.get("/")
async def home():
    """Serve the main HTML interface"""
    return HTMLResponse(content=HTML_TEMPLATE)

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Handle file upload and return sheet information"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(400, "Only .xlsx files are supported")

    # Save file temporarily
    temp_path = Path(tempfile.mktemp(suffix='.xlsx'))
    content = await file.read()
    temp_path.write_bytes(content)

    try:
        # Get sheet information
        import openpyxl
        wb = openpyxl.load_workbook(temp_path, read_only=True)
        sheets = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheets.append({
                'name': sheet_name,
                'rows': sheet.max_row,
                'cols': sheet.max_column
            })

        wb.close()

        # Store file path for later processing
        file_id = temp_path.name
        temp_files[file_id] = temp_path

        return {
            'file_id': file_id,
            'filename': file.filename,
            'sheets': sheets
        }

    except Exception as e:
        temp_path.unlink()
        raise HTTPException(500, f"Error processing file: {str(e)}")

@app.post("/api/process")
async def process_file(
    file: UploadFile = File(...),
    sheets: str = "",
    range: str = "",
    format: str = "preview"
):
    """Process Excel file with selected options"""

    # Save file temporarily if not already saved
    temp_path = Path(tempfile.mktemp(suffix='.xlsx'))
    content = await file.read()
    temp_path.write_bytes(content)

    try:
        # Parse file
        parser = ExcelParser()
        sheet_list = [s.strip() for s in sheets.split(',') if s.strip()] if sheets else None
        range_str = range.strip() if range else None

        data = parser.parse_file(
            str(temp_path),
            sheets=sheet_list,
            ranges=range_str
        )

        # Format response based on requested format
        formatter = ExcelFormatter()

        # Calculate total cells for size warning
        total_cells = sum(len(sheet.get('cells', [])) for sheet in data.get('sheets', []))

        if format == 'preview':
            # Return structured data for preview
            return {
                'format': 'preview',
                'data': data,
                'total_cells': total_cells
            }
        elif format == 'json':
            return {
                'format': 'json',
                'data': data,
                'content': formatter.to_json(data, pretty=True)
            }
        elif format == 'markdown':
            return {
                'format': 'markdown',
                'content': formatter.to_markdown(data)
            }
        elif format == 'text':
            return {
                'format': 'text',
                'content': formatter.to_simple_text(data)
            }
        else:
            raise HTTPException(400, "Invalid format")

    except Exception as e:
        raise HTTPException(500, f"Error processing file: {str(e)}")
    finally:
        temp_path.unlink()

def convert_to_csv(data):
    """Convert data to CSV format"""
    import csv
    import io

    output = io.StringIO()

    for sheet in data['sheets']:
        # Write sheet name as header
        output.write(f"# Sheet: {sheet['name']}\n")

        # Get all unique row/col combinations
        cells_dict = {}
        max_row = 0
        max_col = 0

        for cell in sheet['cells']:
            row, col = cell['row'], cell['col']
            cells_dict[(row, col)] = cell
            max_row = max(max_row, row)
            max_col = max(max_col, col)

        # Write as CSV
        writer = csv.writer(output)

        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = cells_dict.get((row, col))
                if cell:
                    if cell.get('formula'):
                        row_data.append(cell['formula'])
                    else:
                        row_data.append(cell.get('value', ''))
                else:
                    row_data.append('')
            writer.writerow(row_data)

        output.write('\n')

    return output.getvalue()


@app.post("/api/download")
async def download_file(
    file: UploadFile = File(...),
    sheets: str = "",
    range: str = "",
    format: str = "json"
):
    """Generate and download file in requested format"""

    import traceback

    # Process file
    temp_path = Path(tempfile.mktemp(suffix='.xlsx'))
    content = await file.read()
    temp_path.write_bytes(content)

    try:
        # Parse file with chunk limit for large files
        parser = ExcelParser(chunk_size=1000)  # Increase chunk size for download
        sheet_list = [s.strip() for s in sheets.split(',') if s.strip()] if sheets else None
        range_str = range.strip() if range else None

        print(f"Parsing file with sheets: {sheet_list}, range: {range_str}")

        data = parser.parse_file(
            str(temp_path),
            sheets=sheet_list,
            ranges=range_str
        )

        print(f"Data parsed successfully. Format requested: {format}")

        # Check total cells to warn about large files
        total_cells = sum(len(sheet.get('cells', [])) for sheet in data.get('sheets', []))
        if total_cells > 10000:
            print(f"Warning: Large dataset with {total_cells} cells")

        # Format output
        formatter = ExcelFormatter()

        if format == 'json':
            output = formatter.to_json(data, pretty=True)
            media_type = "application/json"
            filename = f"{Path(file.filename).stem}_output.json"
        elif format == 'markdown':
            output = formatter.to_markdown(data)
            media_type = "text/markdown"
            filename = f"{Path(file.filename).stem}_output.md"
        elif format == 'text':
            output = formatter.to_simple_text(data)
            media_type = "text/plain"
            filename = f"{Path(file.filename).stem}_output.txt"
        elif format == 'csv':
            # Convert to CSV format
            output = convert_to_csv(data)
            media_type = "text/csv"
            filename = f"{Path(file.filename).stem}_output.csv"
        else:
            raise HTTPException(400, "Invalid format")

        print(f"Output generated. Length: {len(output)} characters")
        print(f"Media type: {media_type}, Filename: {filename}")

        # Clean up input file
        temp_path.unlink()

        # Return file content using StreamingResponse
        import io
        output_bytes = output.encode('utf-8')

        # Add content-length header for progress tracking
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(output_bytes))
        }

        # For very large files, consider chunking
        if len(output_bytes) > 10 * 1024 * 1024:  # 10MB
            print(f"Warning: Large output file ({len(output_bytes) / 1024 / 1024:.1f} MB)")

        return StreamingResponse(
            io.BytesIO(output_bytes),
            media_type=media_type,
            headers=headers
        )

    except Exception as e:
        print(f"Error in download_file: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        print(f"Traceback: {traceback.format_exc()}")
        if 'temp_path' in locals() and temp_path.exists():
            temp_path.unlink()
        raise HTTPException(500, f"Error generating download: {str(e)}")

# Cleanup old temp files periodically
import asyncio
import time

async def cleanup_temp_files():
    """Clean up old temporary files"""
    while True:
        await asyncio.sleep(3600)  # Run every hour

        current_time = time.time()
        for file_id, path in list(temp_files.items()):
            if path.exists():
                # Remove files older than 1 hour
                if current_time - path.stat().st_mtime > 3600:
                    path.unlink()
                    del temp_files[file_id]

@app.on_event("startup")
async def startup_event():
    """Start background tasks"""
    asyncio.create_task(cleanup_temp_files())

def run_server(host="0.0.0.0", port=8080):
    """Run the web server"""
    print(f"""
    ╔═══════════════════════════════════════════════╗
    ║       ExcelLLM Web Interface Starting...      ║
    ╚═══════════════════════════════════════════════╝

    🌐 Open your browser at: http://localhost:{port}

    📁 Drag & drop Excel files or click to browse
    📊 Select sheets and ranges interactively
    💾 Download in multiple formats

    Press Ctrl+C to stop the server
    """)

    uvicorn.run(app, host=host, port=port)

if __name__ == "__main__":
    run_server()
